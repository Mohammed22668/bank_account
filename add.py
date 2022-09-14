from openpyxl import Workbook , load_workbook

def add_data(list_data):
   ws = load_workbook('work.xlsx')
   wa = ws.active
   wa.append(list_data)
   ws.save("work.xlsx")
   


#def auto_id():
 #  ws = load_workbook('work.xlsx')
 #  wa = ws.active
  # id_list=[]
 #  for row in range(1,len(wa['A'])+1):
      # print(wa['A'+str(row)].value)
    #  id_list.append(int(wa['A'+str(row)].value))
      # print(wa['A'+str(row)].value , type(wa['A'+str(row)].value))
   #return id_list   



<<<<<<< HEAD
=======
#print ("Hello new_word")

print ("anythink")
>>>>>>> 79cc4f16b33fc7b8e7b6c72e723cf61dc977544f


print("Im mohammed this is last update")